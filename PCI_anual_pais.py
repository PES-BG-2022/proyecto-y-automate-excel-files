# ----------------------------------------------
# Carga de librerías
# ----------------------------------------------
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, LineChart, Reference
import string


# ----------------------------------------------
# Carga datos desde fuente
# ----------------------------------------------

# Carga de los datos originales
df = pd.read_excel("https://api.worldbank.org/v2/es/indicator/FP.CPI.TOTL.ZG?downloadformat=excel", sheet_name= "Data", header= 3)


# ----------------------------------------------
# Filtro 1: Información a considerar
# ----------------------------------------------

# Datos en un período seleccionado: 2000 - 2020
df_selec_anio = df.drop(columns=['Country Code', 'Indicator Name', 'Indicator Code', '1960', '1961', '1962', '1963', '1964', '1965', '1966', '1967', '1968', '1969', '1970', '1971', '1972', '1973', '1974', '1975', '1976', '1977', '1978', '1979', '1980', '1981', '1982', '1983', '1984', '1985', '1986', '1987', '1988', '1989', '1990', '1991', '1992', '1993', '1994', '1995', '1996', '1997', '1998', '1999'])
anio = df_selec_anio.round(2)


# ----------------------------------------------
# Filtro 2: Información a considerar
# ----------------------------------------------

# Datos para un grupo de países seleccionados:
df_selec_pais = anio

lista_paises = ['Estados Unidos', 'Canadá', 'México', 'Costa Rica', 'Guatemala', 'Honduras', 'Nicaragua', 'Panamá', 'El Salvador', 'República Dominicana', 'Uruguay', 'Bolivia', 'Brasil', 'Chile', 'Colombia', 'Perú', 'Paraguay', 
]
filtro = df_selec_pais["Country Name"].apply(lambda pais: pais in lista_paises)
salida = df_selec_pais.loc[filtro, :]
salida.style.set_table_styles([{"selector":"thead","props":"background-color:black; color:white;"},
                                {"selector":"th.row_heading", "props": [("background-color", "gray"), ("color", "white"),
                                          ("border", "3px solid black"), ("font-size", "1.2rem"), ("font-style", "italic")]},]).set_precision(2)




# ----------------------------------------------
# Generar Estadística Descriptiva
# ----------------------------------------------
est_descrip = salida.describe()
est_descrip.style.set_table_styles([{"selector":"thead","props":"background-color:black; color:white;"},
                                {"selector":"th.row_heading", "props": [("background-color", "gray"), ("color", "white"),
                                          ("border", "3px solid black"), ("font-size", "1.2rem"), ("font-style", "italic")]},]).set_precision(2)




# ----------------------------------------------
# Generando salido Microsoft Excel Automatizado
# ----------------------------------------------

grafica_1 = salida.T
grafica_1

nombre = "Reporte_2020"

# Creando un excel en blanco
data_null={}
df_null=pd.DataFrame(data_null)
df_null.to_excel(f"CPI_{nombre}.xlsx", index = False)

# ----------------------------------------------
# Agregando pestañas al Excel
def automatizar_excel(nombre_archivo):

    nombre = nombre_archivo
    writer = pd.ExcelWriter(f"CPI_{nombre}.xlsx")

    salida.to_excel(writer, "CPI_Reporte", startcol=1, startrow=1, index=False)    
    
    est_descrip_1.to_excel(writer, "CPI_Reporte_Stats", startcol=1, startrow=1, index=True)

    grafica_1.to_excel(writer, "datos_graf", startcol=0, startrow=0, index=True)         

    writer.save()

    wb = load_workbook(f"CPI_{nombre}.xlsx")
    pestaña = wb["datos_graf"].sheet_state = 'hidden'   

    return

automatizar_excel(nombre)

# ----------------------------------------------

def ocutlar(nombre_archivo):
    
    nombre = nombre_archivo
    wb = load_workbook(f"CPI_{nombre}.xlsx")
    pestaña = wb["datos_graf"].sheet_state = 'hidden'

    wb.save(f"CPI_{nombre}.xlsx")

    return

ocutlar(nombre)

# ----------------------------------------------

#Genera la gráfica del comportamiento de la inflación anual
#para los países seleccionados del reporte del Banco Mundial
def graficar():
    
    wb = load_workbook(f"CPI_{nombre}.xlsx")
    pestaña = wb["datos_graf"]    

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    barchart = LineChart()    
    data = Reference(pestaña, min_col=min_col, max_col=max_col-5, min_row=min_fila, max_row=max_fila+4)
    categorias = Reference(pestaña, min_col=min_col-1, max_col=min_col-1, min_row=min_fila+1, max_row=max_fila+4)

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)
    
    pestaña_2 = wb["CPI_Reporte"]
    pestaña_2.add_chart(barchart, "B24")
    barchart.title = "CPI: Países Seleccionados"
    barchart.style = 2
    barchart.height = 15
    barchart.width = 20

    pestaña_2["B23"] = "Series de Tiempo"
    pestaña_2["B23"].font = Font("Arial", bold=True, size= 20)
    
    wb.save(f"CPI_{nombre}.xlsx")

graficar()



# ----------------------------------------------
# --
# ----------------------------------------------
